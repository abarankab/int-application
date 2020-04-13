import openpyxl as op
import os, django

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "school.settings")
django.setup()

from school_app.models import Student, OlympData

def parse_table(path):
    wb = op.load_workbook(path)
    ws = wb.worksheets[0]

    if path.__contains__('10kl'):
        MIN_COL = 5
    else:
        MIN_COL = 4

    MAX_COL = ws.max_column + 1

    MIN_ROW = 3
    MAX_ROW = ws.max_row + 1

    subjects = []
    dates = []

    for col in range(MIN_COL, MAX_COL):
        subjects.append(ws.cell(row=1, column=col).value)

    for col in range(MIN_COL, MAX_COL):
        if ws.cell(row=2, column=col).value is None:
            print(path, col)
        dates.append(ws.cell(row=2, column=col).value)

    for row in range(MIN_ROW, MAX_ROW):
        id = ws.cell(row=row, column=1).value

        new_student = Student(id=id)
        new_student.save()

        for col in range(MIN_COL, MAX_COL):
            new_olymp_data = OlympData(student_reference=new_student,
                                       subject=subjects[col - MIN_COL],
                                       date=dates[col - MIN_COL],
                                       login=ws.cell(row=row, column=col).value)
            new_olymp_data.save()

olymp_tables = list(map(lambda s: "./olymp-logins/" + s, os.listdir("./olymp-logins")))
for path in olymp_tables:
    parse_table(path)
