import openpyxl as op
import os, django
import matplotlib.pyplot as plot

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "school.settings")
django.setup()

from school_app.models import Subject, Student, Result

print("Какой worksheet парсить?")
ind = int(input()) - 1

wb = op.load_workbook("table.xlsx", data_only=True)
ws = wb.worksheets[ind]

results = []

for i in range(1, 10000):
    A = 'B' + str(i + 1)
    print(A)
    print(ws[A])

    if ws[A].value == None:
        break

    results.append(ws[A].value)

results.sort(reverse=True)

plot.hist(results, 150, normed=1, facecolor='green', alpha=0.75)
plot.show()