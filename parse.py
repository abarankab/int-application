import openpyxl as op
import os, django

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "school.settings")
django.setup()

from school_app.models import Subject, Student, Result

extraordinary_students = []  # Поступают на несколько профилей

def parse_worksheet(ind):
    wb = op.load_workbook("table2019.xlsx")
    ws = wb.worksheets[ind]
    grade = 10

    raw_pattern = ws['C4':'O4']
    pattern = []
    for cell in raw_pattern[0]:
        if cell.value == None:
            break
        pattern.append(cell.value)

    people = []

    for current_subject in pattern:
        exists = Subject.objects.filter(name=current_subject, task_grade=grade, profile_id=ind).count() == 1
        if not exists:
            order_id = ind * 10 if current_subject != "приглашение на собеседование" else ind * 10 + 1
            new_subject = Subject(name=current_subject, task_grade=grade, order_id=order_id, profile_id=ind)
            new_subject.save()

    print(grade)
    print(pattern)

    for i in range(5, 10000):
        A = 'B' + str(i)
        C = 'C' + str(i)
        O = 'O' + str(i)

        if ws[A].value == None:
            break

        raw_row = ws[C:O]
        current_row = []
        for cell in raw_row[0]:
            if cell.number_format == "0%":
                current_row.append(str(int(cell.value * 100)) + '%')
            else:
                current_row.append(cell.value)
        current_row = current_row[:len(pattern)]
        current_id = str(ws[A].value)
        people.append([current_id, current_row])

    print(people)

    for person in people:
        if Student.objects.filter(id=person[0]).count() == 0:
            new_student = Student(id=person[0])
            new_student.save()
        else:
            extraordinary_students.append((person[0]))
            print("Collision, {}".format(person[0]))

        for i in range(len(pattern)):
            if pattern[-2] == "итог" and i < len(pattern) - 2:
                '''
                Пропускаем баллы по отдельным заданиям у соцэков
                '''
                continue
            if person[1][i] != None:
                current_student = Student.objects.get(id=person[0])
                current_subject = Subject.objects.get(name=pattern[i], task_grade=grade, profile_id=ind)

                new_result = Result(student_reference=current_student,
                                    subject_reference=current_subject,
                                    score=person[1][i])

                new_result.save()

total_sheets = len(op.load_workbook("table2019.xlsx").sheetnames)
for ind in range(1, total_sheets, 2):
    parse_worksheet(ind)

print("Поправить руками:")
print(*extraordinary_students)
